"""Generates the small, realistic, anonymized (invented) sample spreadsheets
used by tests/test_golden.py, under tests/fixtures/.

Scenario: "Bright Minds Academy", Class 10-A, 3 students, 3 subjects.
  - Physics: two SCORED tests (one Format A subjective test, one Format B
    MCQ test on the "same" underlying topic -- kept as separate topic labels
    per the pipeline's design choice #4) plus one still-pending TRACKING-ONLY
    test (mixed state).
  - Chemistry: TRACKING-ONLY only, no scores at all yet. Includes a student
    (Diya) who is NOT enrolled in Chemistry but appears in the raw tracking
    grid anyway (as a tutor's transcription would), to exercise enrollment
    gating.
  - Maths: on the roster for everyone, but no test file and no tracking grid
    at all -- the "no data" state.

Run this file directly to (re)write the fixtures:
    python tests/generate_fixtures.py
"""
import os

import openpyxl

FIXTURES_DIR = os.path.join(os.path.dirname(__file__), "fixtures")


def _write_sheet(wb, name, rows):
    ws = wb.create_sheet(name)
    for row in rows:
        ws.append(row)
    return ws


def build_roster_csv():
    path = os.path.join(FIXTURES_DIR, "roster.csv")
    with open(path, "w") as f:
        f.write("S.No,Student Name,Class,Subjects\n")
        f.write("1,Aarav Sharma,10-A,\"Physics, Chemistry, Maths\"\n")
        f.write("2,Diya Patel,10-A,\"Physics, Maths\"\n")
        f.write("3,Kabir Singh,10-A,\"Physics, Chemistry, Maths\"\n")
    return path


def build_physics_test1_format_a():
    """Format A workbook: subjective Coulomb's Law test, 4 questions.
    Kabir does not submit at all. Deliberately uses varied header casing to
    exercise the column normalizer."""
    path = os.path.join(FIXTURES_DIR, "physics_test1_format_a.xlsx")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Test_Scores sheet: wide, informational only -- pipeline must ignore it.
    _write_sheet(wb, "Test_Scores", [
        ["Student Name", "Q1", "Q2", "Q3", "Q4", "Total"],
        ["Aarav Sharma", 2, 2, 1, 3, 8],
        ["Diya Patel", 1, 3, 2, 2, 8],
        ["Kabir Singh", "", "", "", "", ""],
    ])

    _write_sheet(wb, "Test_Scores_Unpivot", [
        ["student_name", "Q.No", "SubmittedStatus", "Score"],
        ["Aarav Sharma", "Q1", "Submitted", 2],
        ["Aarav Sharma", "Q2", "Submitted", 2],
        ["Aarav Sharma", "Q3", "Submitted", 1],
        ["Aarav Sharma", "Q4", "Submitted", 3],
        ["Diya Patel", "Q1", "Submitted", 1],
        ["Diya Patel", "Q2", "Submitted", 3],
        ["Diya Patel", "Q3", "Submitted", 2],
        ["Diya Patel", "Q4", "Submitted", 2],
        ["Kabir Singh", "Q1", "Not submitted", None],
        ["Kabir Singh", "Q2", "Not submitted", None],
        ["Kabir Singh", "Q3", "Not submitted", None],
        ["Kabir Singh", "Q4", "Not submitted", None],
    ])

    # All 4 questions share one topic label -- this is a normal (non
    # comprehensive) single-topic test.
    _write_sheet(wb, "CBSE_Typology", [
        ["Q.No", "Topic", "Difficulty", "CBSE Typology", "Marks", "Theory / Numerical", "Paper ID"],
        ["Q1", "Coulomb's Law (Subjective)", "Easy", "Remembering", 2, "Theory", "PHY-T1"],
        ["Q2", "Coulomb's Law (Subjective)", "Medium", "Understanding", 3, "Theory", "PHY-T1"],
        ["Q3", "Coulomb's Law (Subjective)", "Medium", "Applying", 2, "Numerical", "PHY-T1"],
        ["Q4", "Coulomb's Law (Subjective)", "Hard", "Analysing", 3, "Numerical", "PHY-T1"],
    ])

    wb.save(path)
    return path


def build_physics_test2_format_b():
    """Format B workbook: MCQ test on the same underlying material, kept as
    a DIFFERENT topic label ("... (MCQ)") per design choice #4. Uses the
    school's own non-sequential question-bank IDs, and Typology Q.No naming
    that does NOT textually match the Score sheet's column names, to
    exercise the question-ID normalizer."""
    path = os.path.join(FIXTURES_DIR, "physics_test2_format_b.xlsx")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_sheet(wb, "Score", [
        ["student_name", "submitted / not submitted", "total_marks_obt", "Q216", "Q222", "Q310 (i)"],
        ["Aarav Sharma", "Submitted", 3, 1, 0, 2],
        ["Diya Patel", "Submitted", 3, 1, 1, 1],
        ["Kabir Singh", "Not submitted", 0, "", "", ""],
    ])

    # Q.No spelled differently than the Score sheet's column headers on
    # purpose: "Q.216" vs "Q216", "222" vs "Q222", "Q310(I)" vs "Q310 (i)".
    _write_sheet(wb, "Typology", [
        ["Q.No", "Difficulty", "CBSE Typology", "Marks", "Theory / Numerical", "Paper ID"],
        ["Q.216", "Easy", "remembering", 1, "Theory", "PHY-T2"],
        ["222", "Easy", "Understanding/Remembering", 1, "Theory", "PHY-T2"],
        ["Q310(I)", "Medium", "Applying", 2, "Numerical", "PHY-T2"],
    ])
    # Topic isn't a column in this school's Typology sheet layout at all in
    # this fixture -- add it so the parser's Topic lookup succeeds, matching
    # the "kept as its own topic" MCQ label.
    ws = wb["Typology"]
    ws.cell(row=1, column=7, value="Topic")
    ws.cell(row=2, column=7, value="Coulomb's Law (MCQ)")
    ws.cell(row=3, column=7, value="Coulomb's Law (MCQ)")
    ws.cell(row=4, column=7, value="Coulomb's Law (MCQ)")

    wb.save(path)
    return path


def build_physics_tracking_csv():
    """Format C grid: Physics Test 3 still awaiting grading (mixed state:
    2 tests already scored + 1 pending, within the same subject)."""
    path = os.path.join(FIXTURES_DIR, "physics_tracking.csv")
    with open(path, "w") as f:
        f.write("S.No,Student Name,Test 3\n")
        f.write("1,Aarav Sharma,Submitted\n")
        f.write("2,Diya Patel,Checked\n")
        f.write("3,Kabir Singh,Not submitted\n")
    return path


def build_chemistry_tracking_csv():
    """Format C grid: Chemistry has no scores at all yet (pure
    tracking-only subject). Diya is NOT enrolled in Chemistry (see roster)
    but appears in this raw transcribed grid anyway -- exercises enrollment
    gating: she must not appear in the final trackingOnly output."""
    path = os.path.join(FIXTURES_DIR, "chemistry_tracking.csv")
    with open(path, "w") as f:
        f.write("S.No,Student Name,Unit Test 1,Unit Test 2\n")
        f.write("1,Aarav Sharma,Submitted,Not submitted\n")
        f.write("2,Diya Patel,Submitted,Submitted\n")
        f.write("3,Kabir Singh,Checked,Submitted\n")
    return path


def build_manifest():
    path = os.path.join(FIXTURES_DIR, "manifest.json")
    import json
    manifest = {
        "school": "Bright Minds Academy",
        "className": "10-A",
        "rosterPath": "roster.csv",
        "attendance": {
            "Aarav Sharma": {"attendancePct": 92.5, "presentPct": 85.0, "onlinePct": 7.5, "absentPct": 7.5},
            "Diya Patel": {"attendancePct": 88.0, "presentPct": 80.0, "onlinePct": 8.0, "absentPct": 12.0},
            "Kabir Singh": {"attendancePct": 75.0, "presentPct": 70.0, "onlinePct": 5.0, "absentPct": 25.0},
        },
        "syllabusProgress": {"Physics": 60.0, "Chemistry": 40.0, "Maths": 55.0},
        "subjects": {
            "Physics": {
                "scoredTests": [
                    {
                        "path": "physics_test1_format_a.xlsx",
                        "format": "A",
                        "test_id": "phy_test1",
                        "test_label": "Physics Test 1 (Subjective)",
                    },
                    {
                        "path": "physics_test2_format_b.xlsx",
                        "format": "B",
                        "test_id": "phy_test2",
                        "test_label": "Physics Test 2 (MCQ)",
                    },
                ],
                "trackingGrid": "physics_tracking.csv",
            },
            "Chemistry": {
                "scoredTests": [],
                "trackingGrid": "chemistry_tracking.csv",
            },
            "Maths": {
                "scoredTests": [],
            },
        },
    }
    with open(path, "w") as f:
        json.dump(manifest, f, indent=2)
    return path


def main():
    os.makedirs(FIXTURES_DIR, exist_ok=True)
    build_roster_csv()
    build_physics_test1_format_a()
    build_physics_test2_format_b()
    build_physics_tracking_csv()
    build_chemistry_tracking_csv()
    build_manifest()
    print("Fixtures written to", FIXTURES_DIR)


if __name__ == "__main__":
    main()
